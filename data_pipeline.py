import pandas as pd
import glob
import os
import logging
from datetime import datetime
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class KitokoDataPipeline:
    def __init__(self, data_dir: str):
        self.data_dir = data_dir
        self.campaign_output_file = os.path.join(self.data_dir, 'kitoko project files', 'Kitoko_Campaign_Data.xlsx')
        self.order_details_file = os.path.join(self.data_dir, 'kitoko project files', 'Kitoko order details.xlsx')

    def merge_campaign_data(self):
        """Merges all daily CSV files into a single master campaign Excel file."""
        logging.info("Starting campaign data merge...")
        all_files = glob.glob(os.path.join(self.data_dir, "*.csv"))
        
        if not all_files:
            logging.warning("No CSV files found in the data directory.")
            return None

        dfs = []
        for file in all_files:
            base_name = os.path.basename(file)
            name_without_ext = os.path.splitext(base_name)[0]

            try:
                report_date = datetime.strptime(name_without_ext, "%d %b %Y").date()
            except ValueError:
                try:
                    report_date = datetime.strptime(name_without_ext, "%d %B %Y").date()
                except ValueError:
                    logging.warning(f"Skipping file '{file}' due to incorrect filename format.")
                    continue

            try:
                df = pd.read_csv(file)
                df['Report Date'] = pd.to_datetime(report_date)
                dfs.append(df)
            except Exception as e:
                logging.error(f"Error reading {file}: {e}")

        if dfs:
            final_df = pd.concat(dfs, ignore_index=True)
            
            # Sort newest first
            final_df = final_df.sort_values(by='Report Date', ascending=False)
            
            # Convert back to string format for Excel representation if needed, 
            # though keeping as datetime is generally better for Excel.
            final_df['Report Date'] = final_df['Report Date'].dt.strftime('%d %b %Y')
            
            # Reorder columns to put 'Report Date' first
            cols = ['Report Date'] + [col for col in final_df.columns if col != 'Report Date']
            final_df = final_df[cols]

            # Save
            if os.path.exists(self.campaign_output_file):
                logging.info(f"Existing file '{self.campaign_output_file}' will be overwritten.")
            
            final_df.to_excel(self.campaign_output_file, index=False)
            logging.info(f"Successfully merged {len(dfs)} files and saved to '{self.campaign_output_file}'!")
            return final_df
        else:
            logging.error("No valid dataframes were created from the CSVs.")
            return None

    def segment_label(self, score: str) -> str:
        """Helper to assign segment based on RFM string score."""
        if score == '555':
            return 'Top Performer'
        elif score.startswith('5'):
            return 'Recent Favorite'
        elif score.endswith('5'):
            return 'High Value'
        elif score.startswith('1'):
            return 'At Risk'
        else:
            return 'Mid Performer'

    def compute_rfm_scores(self, reference_date='2025-05-15'):
        """Computes RFM scores and segments from order details, and saves it as a new sheet."""
        logging.info("Starting RFM calculation...")
        try:
            df = pd.read_excel(self.order_details_file, sheet_name='Order Details')
            logging.info("Successfully loaded Order Details.")
        except Exception as e:
            logging.error(f"Failed to load Excel file '{self.order_details_file}': {e}")
            return None

        # Clean Dates
        if 'purchase-date' not in df.columns:
            logging.error("Column 'purchase-date' not found!")
            return None
            
        df['purchase-date'] = pd.to_datetime(df['purchase-date'].astype(str).str[:10], errors='coerce')
        df = df.dropna(subset=['purchase-date'])

        # Clean numeric fields
        df['quantity'] = pd.to_numeric(df.get('quantity', 1), errors='coerce').fillna(1).astype(int)
        
        if 'item-price' not in df.columns:
            logging.error("Column 'item-price' not found!")
            return None
            
        df['item-price'] = pd.to_numeric(df['item-price'], errors='coerce').fillna(0)
        df['total_value'] = df['item-price']

        # Drop missing ASINs
        if 'asin' not in df.columns:
            logging.error("Column 'asin' not found!")
            return None
        df = df.dropna(subset=['asin'])

        # Group by ASIN for RFM
        ref_date = pd.Timestamp(reference_date)
        rfm_df = df.groupby('asin').agg(
            Recency=('purchase-date', lambda x: (ref_date - x.max()).days),
            Frequency=('asin', 'count'),
            Monetary=('total_value', 'sum')
        ).reset_index()

        # Compute RFM Scores (1-5, 5 is best)
        rfm_df['R_Score'] = pd.qcut(rfm_df['Recency'].rank(method='first'), 5, labels=[5, 4, 3, 2, 1])
        rfm_df['F_Score'] = pd.qcut(rfm_df['Frequency'].rank(method='first'), 5, labels=[1, 2, 3, 4, 5])
        rfm_df['M_Score'] = pd.qcut(rfm_df['Monetary'].rank(method='first'), 5, labels=[1, 2, 3, 4, 5])

        # Concatenate and determine Segment
        rfm_df['RFM_Score'] = rfm_df['R_Score'].astype(str) + rfm_df['F_Score'].astype(str) + rfm_df['M_Score'].astype(str)
        rfm_df['Segment'] = rfm_df['RFM_Score'].apply(self.segment_label)

        # Sort
        rfm_df = rfm_df.sort_values(by='Monetary', ascending=False)
        logging.info(f"RFM DataFrame created. Sample:\n{rfm_df.head()}")

        # Save back to Excel
        sheet_output = 'RFM_Scores_new'
        try:
            book = load_workbook(self.order_details_file)
            if sheet_output in book.sheetnames:
                del book[sheet_output]
                # It's safer to not save immediately, let the writer handle it or save once.
            
            with pd.ExcelWriter(self.order_details_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                rfm_df.to_excel(writer, sheet_name=sheet_output, index=False)
            logging.info(f"RFM analysis saved to sheet '{sheet_output}' in {self.order_details_file}")
        except Exception as e:
            logging.error(f"Error saving RFM data to Excel: {e}")
            
        return rfm_df

if __name__ == "__main__":
    # Dynamically resolve directory to make the project portable
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    DATA_DIR = os.path.dirname(SCRIPT_DIR)
    
    pipeline = KitokoDataPipeline(DATA_DIR)
    
    # 1. Merge Campaign Data
    pipeline.merge_campaign_data()
    
    # 2. Compute RFM Scores
    pipeline.compute_rfm_scores()
